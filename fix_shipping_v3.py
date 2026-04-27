import openpyxl, sys, io, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

src = r"C:\RenStudio\case\testcase\shopee\test_v4.xlsx"
dst = r"C:\RenStudio\case\testcase\shopee\test_v5.xlsx"
shutil.copy(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb.active
headers = [c.value for c in ws[1]]
def col(name): return headers.index(name) + 1

# 把「套用全店運費」改 no，讓商品用自帶運費
target = "套用全店運費"
c = col(target)
for row in range(2, ws.max_row + 1):
    ws.cell(row=row, column=c).value = "no"
    print(f"  row {row}: {target} = no")

wb.save(dst)
print(f"\n✅ 套用全店運費 → no，存檔 → test_v5.xlsx")
print("(商品自帶運費：7-11/萊爾富 = 60、郵寄掛號 = 35 維持不變)")
