import openpyxl, sys, io
from openpyxl.utils import get_column_letter, column_index_from_string
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\RenStudio\case\testcase\shopee\test_v6.xlsx", read_only=True)
ws = wb[wb.sheetnames[0]]

bc_idx = column_index_from_string("BC")
print(f"BC 欄 = 第 {bc_idx} 欄")

# 第一列（headers）
header_at_bc = ws.cell(row=1, column=bc_idx).value
print(f"BC1 (header) = {header_at_bc!r}")

# 找屬性欄位1 真正在第幾欄
headers = [c.value for c in ws[1]]
attr1_idx = headers.index("屬性欄位1") + 1 if "屬性欄位1" in headers else -1
print(f"屬性欄位1 真實位置 = 第 {attr1_idx} 欄 = {get_column_letter(attr1_idx)}")

# 看第 31 列（ZTE WiFi 路由器）目前 BC 值
print(f"\n第 31 列 商品名稱: {ws.cell(row=31, column=headers.index('商品名稱')+1).value}")
print(f"第 31 列 類別: {ws.cell(row=31, column=headers.index('類別')+1).value}")
print(f"第 31 列 BC 值: {ws.cell(row=31, column=bc_idx).value!r}")
