import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\RenStudio\case\testcase\shopee\test_v6.xlsx", read_only=True)
print(f"工作表清單: {wb.sheetnames}")

ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
headers = list(rows[0])
data = rows[1:]
print(f"第一個工作表「{ws.title}」有 {len(data)} 筆")

# 找重要欄位
def col(name):
    try: return headers.index(name)
    except: return -1

ni = col("商品名稱")
ci = col("類別")
attr1_i = col("屬性欄位1")
bsmi_i = col("商品檢驗字號")

print(f"\n「商品檢驗字號」欄位 index = {bsmi_i}")
print(f"「屬性欄位1」欄位 index = {attr1_i}")

# 錯誤對應的列：2, 4, 12, 13, 30, 31, 32, 34
target = [2, 4, 12, 13, 30, 31, 32, 34]
print(f"\n=== 錯誤行的商品內容 ===")
for r in target:
    idx = r - 2  # 工作表第 2 列 = data[0]
    if 0 <= idx < len(data):
        row = data[idx]
        print(f"\n[列 {r}] 類別 {row[ci]}")
        print(f"  商品名稱: {row[ni]}")
        print(f"  屬性欄位1: {row[attr1_i]!r}")
        if bsmi_i >= 0:
            print(f"  商品檢驗字號: {row[bsmi_i]!r}")
