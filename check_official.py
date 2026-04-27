"""看官方範本的屬性欄位1 內容，找格式範例"""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\RenStudio\case\testcase\shopee\official_template.xlsx", data_only=True)
print(f"工作表清單: {wb.sheetnames}\n")

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"=== 工作表「{sn}」: {ws.max_row} 列 × {ws.max_column} 欄 ===")
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    headers = list(rows[0])
    # 找屬性欄位
    attr_cols = [(i, h) for i, h in enumerate(headers) if h and "屬性" in str(h)]
    print(f"  屬性相關欄位: {attr_cols[:5]}")

    # 看資料列
    if len(rows) > 1:
        for ri, row in enumerate(rows[1:6], start=2):
            for ci, h in attr_cols:
                v = row[ci] if ci < len(row) else None
                if v not in (None, ""):
                    print(f"  列 {ri} | {h}: {v!r}")
    print()
