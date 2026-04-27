import openpyxl, sys, io, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

src = r"C:\RenStudio\case\testcase\shopee\test_v3.xlsx"
dst = r"C:\RenStudio\case\testcase\shopee\test_v4.xlsx"
shutil.copy(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb.active
headers = [c.value for c in ws[1]]
def col(name): return headers.index(name) + 1

# 收款方式 → 運送方式 + 運費金額
mapping = {
    "接受輕鬆付7-11取貨付款": ("7-11取貨付款", 60),
    "接受輕鬆付萊爾富取貨付款": ("萊爾富取貨付款", 60),
    "接受輕鬆付全家取貨付款": ("全家取貨付款", 60),
    "接受輕鬆付郵局貨到付款": ("郵局貨到付款", 80),
}
prepay_triggers = ["接受輕鬆付現金付款", "接受輕鬆付信用卡一次付清"]
prepay_target = ("郵寄掛號", 35)

fixed = 0
for row in range(2, ws.max_row + 1):
    for pay_col, (ship_col, fee) in mapping.items():
        if str(ws.cell(row=row, column=col(pay_col)).value).lower() == "yes":
            cell = ws.cell(row=row, column=col(ship_col))
            cell.value = fee  # 直接填數字，覆蓋之前錯填的 "yes"
            fixed += 1
            print(f"  row {row}: {ship_col} = {fee}")
    if any(str(ws.cell(row=row, column=col(t)).value).lower() == "yes" for t in prepay_triggers):
        ship_col, fee = prepay_target
        ws.cell(row=row, column=col(ship_col)).value = fee
        fixed += 1
        print(f"  row {row}: {ship_col} = {fee}")

wb.save(dst)
print(f"\n✅ 共修正 {fixed} 個運費欄位 → test_v4.xlsx")
