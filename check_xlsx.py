import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\RenStudio\case\testcase\shopee\拍賣蝦皮拍賣商品_匯出商品資料_20260421_024920.xlsx", read_only=True, data_only=True)
ws = wb.active
print(f"Sheet name: {ws.title}")
print(f"Max col: {ws.max_column}")

rows = ws.iter_rows(values_only=True)
headers = next(rows)
print(f"\nTotal columns: {len(headers)}")
print("\n=== Headers ===")
for i, h in enumerate(headers):
    print(f"  [{i:3d}] {h}")

data_rows = list(rows)
print(f"\nTotal data rows: {len(data_rows)}")

if data_rows:
    print("\n=== First product (full) ===")
    for i, (h, v) in enumerate(zip(headers, data_rows[0])):
        v_short = str(v)[:80] if v is not None else "<empty>"
        print(f"  [{i:3d}] {h}: {v_short}")
