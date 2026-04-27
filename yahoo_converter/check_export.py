"""檢查使用者匯出的 yahoo_ready.xlsx，看列 15、868、902、903 是什麼"""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\Users\boren\Downloads\yahoo_ready.xlsx", read_only=True)
print(f"工作表: {wb.sheetnames}")
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
headers = list(rows[0])
data = rows[1:]
print(f"工作表「{ws.title}」共 {len(data)} 筆\n")

ni = headers.index("商品名稱")
ci = headers.index("類別")
attr1_i = headers.index("屬性欄位1")

for r in [15, 868, 902, 903]:
    idx = r - 2
    if 0 <= idx < len(data):
        row = data[idx]
        print(f"[列 {r}] 類別 {row[ci]}")
        print(f"  商品: {row[ni]}")
        print(f"  屬性欄位1: {row[attr1_i]!r}")
        print()
