"""
Yahoo 拍賣大量上架轉檔核心邏輯（v2 — 依官方範本說明重寫）

修正規則：
1. 收款方式 vs 運送方式對應：接受輕鬆付X = yes → 對應運送方式必須有運費
2. 運送方式型別：用到的填整數運費（0~500）、不用的留白（不是 yes/no）
3. 套用全店運費 → 預設改 no（避免依賴 Yahoo 後台設定）
4. 屬性欄位1 為空 → 該商品分流到「需人工處理」（屬性欄位內容依類別代碼決定，無法純靠 Excel 推導）
5. 商品數量空白 + 無多規格 → 分流

注意：官方範本沒有「商品檢驗字號」獨立欄位，BSMI/NCC 是某些類別的「屬性欄位1」內容。
"""
from dataclasses import dataclass, field
from typing import Callable
import openpyxl
from openpyxl import Workbook


# 應施檢驗類別關鍵字（用於提示「屬性欄位1 可能要填 BSMI」）
BSMI_HINT_KEYWORDS = [
    "電池", "鋰電", "鉛酸", "充電器", "變壓器", "電源供應器",
    "充電線", "傳輸線", "行動電源",
    "燈", "LED", "燈泡", "燈座", "檯燈", "夜燈",
    "電鑽", "電起子", "電動工具",
    "吸塵器", "掃地機",
    "風扇", "暖爐", "電暖", "除濕", "加濕",
    "電鍋", "電熱", "烘碗",
    "玩具", "童車", "兒童", "嬰兒", "奶嘴",
    "安全帽",
]

# 通訊類關鍵字（用於提示「屬性欄位1 可能要填 NCC」）
NCC_HINT_KEYWORDS = [
    "路由器", "WiFi", "Wi-Fi", "分享器", "AP", "中繼器",
    "對講機", "無線電", "藍牙喇叭",
    "手機", "平板",
]


@dataclass
class ConvertConfig:
    fee_711: int = 60
    fee_familymart: int = 60
    fee_hilife: int = 60
    fee_post_cod: int = 80
    fee_mail: int = 35
    fee_home: int = 100
    use_store_shipping: bool = False
    bsmi_keywords: list = field(default_factory=lambda: list(BSMI_HINT_KEYWORDS))
    ncc_keywords: list = field(default_factory=lambda: list(NCC_HINT_KEYWORDS))
    # 類別屬性樣板：{類別代碼: 屬性代碼字串}（由使用者在 GUI 中設定）
    category_attr_templates: dict = field(default_factory=dict)


@dataclass
class FixLogEntry:
    row_index: int
    product_name: str
    field: str
    old_value: object
    new_value: object
    rule: str


@dataclass
class ConversionResult:
    headers: list
    uploadable: list
    needs_review: list
    fix_log: list
    stats: dict


PAY_TO_SHIP_MAP = {
    "接受輕鬆付7-11取貨付款": "7-11取貨付款",
    "接受輕鬆付萊爾富取貨付款": "萊爾富取貨付款",
    "接受輕鬆付全家取貨付款": "全家取貨付款",
    "接受輕鬆付郵局貨到付款": "郵局貨到付款",
}
PREPAY_TRIGGERS = ["接受輕鬆付現金付款", "接受輕鬆付信用卡一次付清"]
PREPAY_DEFAULT_SHIP = "郵寄掛號"


def _is_yes(v) -> bool:
    return str(v).strip().lower() == "yes"


def _fee_for(ship_field: str, cfg: ConvertConfig) -> int:
    return {
        "7-11取貨付款": cfg.fee_711,
        "萊爾富取貨付款": cfg.fee_hilife,
        "全家取貨付款": cfg.fee_familymart,
        "郵局貨到付款": cfg.fee_post_cod,
        "郵寄掛號": cfg.fee_mail,
        "宅配": cfg.fee_home,
    }.get(ship_field, 0)


def hint_attr_type(name: str, cfg: ConvertConfig) -> str:
    """根據商品名稱推測屬性欄位1 可能需要填什麼類型（提示用，非絕對）"""
    text = str(name or "")
    for kw in cfg.ncc_keywords:
        if kw in text:
            return "NCC型式認證碼"
    for kw in cfg.bsmi_keywords:
        if kw in text:
            return "BSMI商品檢驗字號"
    return "依類別需求"


def convert(
    input_path: str,
    cfg: ConvertConfig,
    progress_cb: Callable[[int, int], None] = None,
) -> ConversionResult:
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    data = [r for r in rows_iter]
    total = len(data)

    h_index = {h: i for i, h in enumerate(headers)}

    def col(name: str) -> int:
        return h_index.get(name, -1)

    name_i = col("商品名稱")
    qty_i = col("商品數量")
    cat_i = col("類別")
    apply_store_i = col("套用全店運費")
    attr1_i = col("屬性欄位1")
    spec1_i = col("第一層規格:名稱")

    # 所有運送方式欄位（要清掉錯誤的 yes/no 字串）
    SHIP_FIELDS = [
        "郵寄掛號", "7-11取貨付款", "萊爾富取貨付款", "全家取貨付款",
        "郵局貨到付款", "宅配", "低溫寄送", "離島寄送", "跨國寄送",
        "面交/自取/不寄送", "大型/超重物品寄送", "7-11取貨不付款", "取貨付款",
    ]

    uploadable = []
    needs_review = []
    fix_log: list[FixLogEntry] = []
    stats = {
        "total": total,
        "uploadable": 0,
        "review_attr": 0,
        "review_qty": 0,
        "review_other": 0,
        "fixes": 0,
    }

    for idx, raw_row in enumerate(data):
        if progress_cb:
            progress_cb(idx + 1, total)
        row = list(raw_row)
        name = row[name_i] if name_i >= 0 else ""
        category = str(row[cat_i]) if cat_i >= 0 and row[cat_i] is not None else ""

        review_reasons = []

        # 規則 3：套用全店運費 → no
        if apply_store_i >= 0 and not cfg.use_store_shipping:
            old = row[apply_store_i]
            if str(old).strip().lower() != "no":
                _log(fix_log, idx, name, "套用全店運費", old, "no", "改為商品自帶運費")
                row[apply_store_i] = "no"

        # 規則 2 (Step A)：清掉所有運送方式欄位裡的非數字值（如 yes / no）
        for sf in SHIP_FIELDS:
            si = col(sf)
            if si < 0:
                continue
            v = row[si]
            if v is None or v == "":
                continue
            if isinstance(v, (int, float)):
                continue
            # 非數字 → 視為「未設定」，留空
            _log(fix_log, idx, name, sf, v, "(留空)", "運送方式應為運費數字或留空")
            row[si] = None

        # 規則 1+2 (Step B)：依收款方式設運費
        for pay_field, ship_field in PAY_TO_SHIP_MAP.items():
            pi = col(pay_field)
            si = col(ship_field)
            if pi < 0 or si < 0:
                continue
            if _is_yes(row[pi]):
                fee = _fee_for(ship_field, cfg)
                old = row[si]
                if old != fee:
                    _log(fix_log, idx, name, ship_field, old, fee, f"配對 {pay_field}")
                    row[si] = fee

        if any(_is_yes(row[col(t)]) for t in PREPAY_TRIGGERS if col(t) >= 0):
            si = col(PREPAY_DEFAULT_SHIP)
            fee = cfg.fee_mail
            if si >= 0 and row[si] != fee:
                _log(fix_log, idx, name, PREPAY_DEFAULT_SHIP, row[si], fee, "預付類收款需至少一種預付運送")
                row[si] = fee

        # 規則 4：屬性欄位1（保守邏輯）
        # 客戶有義務為每個類別決定要填字號還是 (none)。工具不擅自代客戶決定。
        # 沒填樣板 → 視為「未確認」→ 分流到需人工處理。
        if attr1_i >= 0:
            template = cfg.category_attr_templates.get(category)
            if template is not None and template.strip():
                if template.strip().lower() in ("(none)", "(空)", "(skip)"):
                    if row[attr1_i] not in (None, ""):
                        _log(fix_log, idx, name, "屬性欄位1", row[attr1_i], "(留空)", f"類別 {category} 標示無必填屬性")
                        row[attr1_i] = None
                else:
                    if row[attr1_i] != template:
                        _log(fix_log, idx, name, "屬性欄位1", row[attr1_i], template, f"套用類別 {category} 樣板")
                        row[attr1_i] = template
            elif row[attr1_i] in (None, ""):
                hint = hint_attr_type(name, cfg)
                review_reasons.append(f"屬性欄位1 未確認（類別 {category} 推測需 {hint}）")

        # 規則 5：商品數量檢查（多規格商品數量本應為空，跳過）
        has_spec = spec1_i >= 0 and row[spec1_i] not in (None, "")
        if not has_spec and qty_i >= 0 and row[qty_i] in (None, ""):
            review_reasons.append("商品數量為空白且無規格")

        # 分流
        if review_reasons:
            needs_review.append(tuple(row) + (" / ".join(review_reasons),))
            r0 = review_reasons[0]
            if "屬性欄位1" in r0 or "推測需要" in r0:
                stats["review_attr"] += 1
            elif "數量" in r0:
                stats["review_qty"] += 1
            else:
                stats["review_other"] += 1
        else:
            uploadable.append(tuple(row))
            stats["uploadable"] += 1

    stats["fixes"] = len(fix_log)
    return ConversionResult(
        headers=headers, uploadable=uploadable, needs_review=needs_review,
        fix_log=fix_log, stats=stats,
    )


def _log(log_list, row_idx, name, field_name, old, new, rule):
    log_list.append(FixLogEntry(
        row_index=row_idx + 2,
        product_name=str(name)[:60] if name else "",
        field=field_name,
        old_value=old,
        new_value=new,
        rule=rule,
    ))


def write_output(result: ConversionResult, output_path: str):
    """
    輸出純淨的商品檔，符合 Yahoo 大量刊登助手規格。
    不在檔案中加任何警告/提醒文字（避免污染商品資料），
    法律提醒一律在 GUI 端呈現。
    """
    wb = Workbook()

    # Sheet 1（必須是商品資料 — Yahoo 工具會讀第一個工作表）
    ws1 = wb.active
    ws1.title = "可上架商品"
    ws1.append(result.headers)
    for r in result.uploadable:
        ws1.append(r)

    # Sheet 2：需人工處理（輔助用，Yahoo 工具不讀）
    ws2 = wb.create_sheet("需人工處理")
    ws2.append(list(result.headers) + ["未通過原因"])
    for r in result.needs_review:
        ws2.append(r)

    # Sheet 3：修正紀錄
    ws3 = wb.create_sheet("修正紀錄")
    ws3.append(["原檔列號", "商品名稱", "欄位", "原值", "新值", "套用規則"])
    for e in result.fix_log:
        ws3.append([e.row_index, e.product_name, e.field, str(e.old_value), str(e.new_value), e.rule])

    wb.save(output_path)


# 提供給 GUI 用：分析檔案的類別分布
def list_categories(input_path: str, cfg: ConvertConfig = None) -> list:
    """回傳 [(類別代碼, 商品數, 範例名稱, 屬性提示)]"""
    if cfg is None:
        cfg = ConvertConfig()
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    ci = headers.index("類別")
    ni = headers.index("商品名稱")

    from collections import Counter, defaultdict
    counter = Counter()
    samples = defaultdict(list)
    for r in rows:
        cat = str(r[ci]) if r[ci] is not None else ""
        if not cat:
            continue
        counter[cat] += 1
        if len(samples[cat]) < 3:
            samples[cat].append(str(r[ni]) if r[ni] else "")

    return [
        (cat, n, samples[cat][0] if samples[cat] else "", hint_attr_type(samples[cat][0] if samples[cat] else "", cfg))
        for cat, n in counter.most_common()
    ]
