"""跑一次完整 1052 筆檔案，驗證 converter.py 沒問題"""
import sys, io, os
sys.path.insert(0, os.path.dirname(__file__))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from converter import convert, write_output, ConvertConfig

INPUT = r"C:\RenStudio\case\testcase\shopee\拍賣蝦皮拍賣商品_匯出商品資料_20260421_024920.xlsx"
OUTPUT = r"C:\RenStudio\case\testcase\shopee\yahoo_converter\output_test.xlsx"

cfg = ConvertConfig()
print(f"輸入檔: {INPUT}")
print(f"使用預設 BSMI 關鍵字: {len(cfg.bsmi_keywords)} 組")
print(f"運費設定: 7-11/全家/萊爾富={cfg.fee_711}, 郵局={cfg.fee_post_cod}, 掛號={cfg.fee_mail}\n")

def progress(i, total):
    if i % 100 == 0 or i == total:
        print(f"  處理中... {i}/{total}", flush=True)

result = convert(INPUT, cfg, progress_cb=progress)

print(f"\n=== 統計 ===")
print(f"  總數:           {result.stats['total']}")
print(f"  ✅ 可上架:       {result.stats['uploadable']}")
print(f"  ⚠️  需人工處理:   {len(result.needs_review)}")
print(f"     - 屬性欄位1空:  {result.stats['review_attr']}")
print(f"     - 數量空白:    {result.stats['review_qty']}")
print(f"     - 其他:        {result.stats['review_other']}")
print(f"  📋 自動修正次數: {result.stats['fixes']}")

write_output(result, OUTPUT)
print(f"\n✅ 輸出 → {OUTPUT}")

# 檢查輸出檔
import openpyxl
wb = openpyxl.load_workbook(OUTPUT, read_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"  工作表「{sn}」: {ws.max_row - 1} 筆資料")
