"""
Yahoo 拍賣大量上架轉檔工具 — GUI
Ren Studio 仁格數位科技
"""
import json
import os
import sys
import threading
from tkinter import filedialog, messagebox, ttk
import customtkinter as ctk

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from converter import (
    convert, write_output, list_categories, hint_attr_type,
    ConvertConfig, BSMI_HINT_KEYWORDS as DEFAULT_BSMI_KEYWORDS,
)


# 推測類型 → 占位符 + 顯示色
HINT_TO_PLACEHOLDER = {
    "NCC型式認證碼": ("NCC型式認證碼-:CCAB12LP3210T0", "#0369a1"),  # 藍
    "BSMI商品檢驗字號": ("商品檢驗字號-:R12345", "#b45309"),        # 橘
}
DEFAULT_HINT_COLOR = "#6b7280"  # 灰


ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


FONT_SECTION = ("Microsoft JhengHei", 17, "bold")     # 區塊標題
FONT_TITLE = ("Microsoft JhengHei", 20, "bold")       # 大標題
FONT_BUTTON = ("Microsoft JhengHei", 14, "bold")      # 按鈕
FONT_BIG_BUTTON = ("Microsoft JhengHei", 16, "bold")  # 主動作大按鈕
FONT_LABEL = ("Microsoft JhengHei", 14)               # 一般文字
FONT_BODY = ("Microsoft JhengHei", 13)                # 表格內容
FONT_HINT = ("Microsoft JhengHei", 12)                # 提示文字
FONT_MONO = ("Consolas", 13)                          # 類別代碼等等寬


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Yahoo 拍賣大量上架轉檔工具 - Ren Studio")
        self.geometry("1280x920")
        self.minsize(1200, 820)

        self.input_path: str = ""
        self.result = None
        self.info_banner = None

        # 主容器：左側流程 + 右側結果
        self.tabview = ctk.CTkTabview(self, width=960, height=700)
        self.tabview.pack(fill="both", expand=True, padx=10, pady=10)

        self.tab_step1 = self.tabview.add("Step 1：選擇檔案")
        self.tab_step2 = self.tabview.add("Step 2：設定參數")
        self.tab_step3 = self.tabview.add("Step 3：預覽結果")

        self._build_step1()
        self._build_step2()
        self._build_step3()

        self.tabview.set("Step 1：選擇檔案")

    # ---------- Step 1 ----------
    def _build_step1(self):
        frame = self.tab_step1
        ctk.CTkLabel(
            frame, text="請選擇從 EC-Mart 大量下載助手匯出的蝦皮商品 Excel",
            font=FONT_TITLE,
        ).pack(pady=(40, 8))
        ctk.CTkLabel(
            frame, text="檔名通常為「拍賣蝦皮拍賣商品_匯出商品資料_yyyymmdd_xxxxxx.xlsx」",
            font=FONT_HINT, text_color="gray",
        ).pack(pady=(0, 24))

        self.lbl_path = ctk.CTkLabel(frame, text="（尚未選擇）", font=FONT_LABEL, wraplength=800)
        self.lbl_path.pack(pady=8)

        ctk.CTkButton(
            frame, text="📁 選擇 Excel 檔...", command=self._pick_file,
            width=260, height=48, font=FONT_BIG_BUTTON,
        ).pack(pady=16)

        self.lbl_step1_info = ctk.CTkLabel(frame, text="", font=FONT_LABEL)
        self.lbl_step1_info.pack(pady=8)

        ctk.CTkButton(
            frame, text="下一步：設定參數 →",
            command=lambda: self.tabview.set("Step 2：設定參數"),
            width=240, height=44, font=FONT_BUTTON,
        ).pack(pady=16)

    def _pick_file(self):
        path = filedialog.askopenfilename(
            title="選擇 Excel 檔",
            filetypes=[("Excel 檔", "*.xlsx"), ("所有檔案", "*.*")],
        )
        if not path:
            return
        self.input_path = path
        self.lbl_path.configure(text=path)
        # 快速統計商品數
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            n = ws.max_row - 1
            self.lbl_step1_info.configure(
                text=f"✅ 偵測到 {n} 筆商品  (工作表：{ws.title})",
                text_color="#1a7f37",
            )
        except Exception as e:
            self.lbl_step1_info.configure(text=f"❌ 讀檔失敗: {e}", text_color="red")

    # ---------- Step 2 ----------
    def _build_step2(self):
        outer = self.tab_step2
        # 內容包 ScrollableFrame 避免擠不下
        frame = ctk.CTkScrollableFrame(outer, fg_color="transparent")
        frame.pack(fill="both", expand=True)

        title_row = ctk.CTkFrame(frame, fg_color="transparent")
        title_row.pack(fill="x", padx=10, pady=(4, 2))
        ctk.CTkLabel(title_row, text="設定轉檔參數", font=FONT_TITLE).pack(side="left")
        ctk.CTkLabel(
            title_row,
            text="⚠️ 這些參數會影響最終上架資料，請確認後再轉檔",
            font=FONT_HINT, text_color="#b45309",
        ).pack(side="left", padx=12)

        # === 上半段：左邊運費 + 右邊樣板按鈕欄 ===
        top_row = ctk.CTkFrame(frame, fg_color="transparent")
        top_row.pack(fill="x", padx=10, pady=2)

        # --- 左：運費 ---
        fee_box = ctk.CTkFrame(top_row)
        fee_box.pack(side="left", fill="both", expand=True, padx=(0, 6))
        ctk.CTkLabel(fee_box, text="📦 運費設定（NT$）", font=FONT_SECTION).grid(row=0, column=0, columnspan=6, sticky="w", padx=8, pady=(6, 2))

        self.fee_vars = {}
        defaults = [
            ("7-11 取貨付款", "7-11", 60),
            ("全家取貨付款", "全家", 60),
            ("萊爾富取貨付款", "萊爾富", 60),
            ("郵局貨到付款", "郵局貨到付款", 80),
            ("郵寄掛號", "郵寄掛號", 35),
            ("宅配", "宅配", 100),
        ]
        for i, (key, short_lbl, default) in enumerate(defaults):
            r = 1 + i // 3
            c = (i % 3) * 2
            ctk.CTkLabel(fee_box, text=short_lbl, font=FONT_LABEL).grid(row=r, column=c, sticky="e", padx=(8, 4), pady=3)
            v = ctk.StringVar(value=str(default))
            self.fee_vars[key] = v
            ctk.CTkEntry(fee_box, textvariable=v, width=80, font=FONT_LABEL).grid(row=r, column=c+1, padx=(0, 6), pady=3)

        self.use_store_var = ctk.BooleanVar(value=False)
        ctk.CTkSwitch(fee_box, text="套用全店運費（除非您已在 Yahoo 後台設好，否則不建議）", variable=self.use_store_var, font=FONT_HINT).grid(row=3, column=0, columnspan=6, sticky="w", padx=8, pady=(4, 6))

        # --- 右：樣板操作按鈕欄（垂直排列）---
        op_col = ctk.CTkFrame(top_row, width=270)
        op_col.pack(side="right", fill="y")
        op_col.pack_propagate(False)

        ctk.CTkLabel(op_col, text="🛠️ 字號樣板操作", font=FONT_SECTION).pack(anchor="w", padx=8, pady=(6, 4))
        ctk.CTkButton(op_col, text="📂 載入檔案類別", command=self._load_categories, height=36, font=FONT_BUTTON).pack(fill="x", padx=8, pady=2)
        ctk.CTkButton(op_col, text="🧪 全填測試字號", command=self._fill_placeholders, height=36, fg_color="#9333ea", font=FONT_BUTTON).pack(fill="x", padx=8, pady=2)
        ctk.CTkButton(op_col, text="💾 儲存字號資料", command=self._save_templates_json, height=36, font=FONT_BUTTON).pack(fill="x", padx=8, pady=2)
        ctk.CTkButton(op_col, text="📥 填入已存在之字號", command=self._load_templates_json, height=36, font=FONT_BUTTON).pack(fill="x", padx=8, pady=2)

        self.lbl_tpl_progress = ctk.CTkLabel(op_col, text="尚未載入類別清單", font=FONT_LABEL, text_color="gray")
        self.lbl_tpl_progress.pack(anchor="w", padx=8, pady=(6, 4))

        # === 下半段：類別字號設定表格 ===
        tpl_box = ctk.CTkFrame(frame)
        tpl_box.pack(fill="both", expand=True, padx=10, pady=4)
        title_box = ctk.CTkFrame(tpl_box, fg_color="transparent")
        title_box.pack(fill="x", padx=8, pady=(6, 2))
        ctk.CTkLabel(title_box, text="🏷️ 各類別字號設定", font=FONT_SECTION).pack(side="left")
        ctk.CTkLabel(
            title_box,
            text="格式：屬性名稱-:值（單選） / 屬性名稱+:值1,值2（複選） / (none)：無必填屬性",
            font=FONT_HINT, text_color="gray",
        ).pack(side="left", padx=10)

        # 表格 header
        header_box = ctk.CTkFrame(tpl_box, fg_color="#e5e7eb", height=34)
        header_box.pack(fill="x", padx=8, pady=(2, 0))
        for text, w, anchor in [
            ("類別代碼", 105, "w"),
            ("商品數", 55, "center"),
            ("推測必填", 130, "w"),
            ("代表商品（點擊看完整品名）", 420, "w"),
            ("字號 / 屬性值（請填）", 360, "w"),
        ]:
            ctk.CTkLabel(header_box, text=text, font=FONT_BODY, width=w, anchor=anchor).pack(side="left", padx=3, pady=4)

        # ScrollableFrame 每個類別一列
        self.tpl_scroll = ctk.CTkScrollableFrame(tpl_box, height=340)
        self.tpl_scroll.pack(fill="both", expand=True, padx=8, pady=(0, 6))

        # 追蹤每個類別的 entry widget
        self.template_entries: dict = {}   # {類別代碼: CTkEntry}
        self.template_metadata: dict = {}  # {類別代碼: (count, hint, sample)}
        self.kw_text = None  # 已不用

        # === 法律確認 ===
        legal_box = ctk.CTkFrame(frame, fg_color="#fef2f2", border_color="#dc2626", border_width=2)
        legal_box.pack(fill="x", padx=10, pady=4)
        ctk.CTkLabel(
            legal_box, text="⚠️ 重要法律聲明（請閱讀後勾選）",
            font=FONT_SECTION, text_color="#dc2626",
        ).pack(padx=8, pady=(6, 2), anchor="w")
        legal_text = (
            "1. 本工具僅處理 Excel 格式轉換，不驗證 BSMI / NCC 字號真實性。\n"
            "2. 工具填入的占位符（R12345、CCAB12LP3210T0）僅供通過匯入驗證，實際上架會被 Yahoo 拒絕。\n"
            "3. 上架前請逐筆確認應施檢驗商品已填真實 BSMI 字號、通訊類商品已填真實 NCC 認證碼。\n"
            "4. 不實標示依《商品檢驗法》可處 20–200 萬罰鍰，責任由賣家承擔。"
        )
        ctk.CTkLabel(legal_box, text=legal_text, font=FONT_HINT, justify="left", text_color="#7f1d1d").pack(padx=8, pady=2, anchor="w")
        self.legal_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            legal_box, text="我已閱讀並同意上述聲明", variable=self.legal_var,
            font=FONT_BUTTON, text_color="#dc2626",
        ).pack(padx=8, pady=(2, 6), anchor="w")

        # === 動作按鈕 ===
        btn_box = ctk.CTkFrame(frame, fg_color="transparent")
        btn_box.pack(fill="x", padx=10, pady=6)
        ctk.CTkButton(btn_box, text="← 上一步", command=lambda: self.tabview.set("Step 1：選擇檔案"), width=110, height=40, font=FONT_BUTTON).pack(side="left")
        ctk.CTkButton(
            btn_box, text="🚀 開始轉檔", command=self._run_convert,
            width=200, height=44, font=FONT_BIG_BUTTON,
        ).pack(side="right")

        self.progress = ctk.CTkProgressBar(frame)
        self.progress.set(0)
        self.progress.pack(fill="x", padx=10, pady=(2, 4))
        self.lbl_progress = ctk.CTkLabel(frame, text="", font=FONT_LABEL)
        self.lbl_progress.pack()

    def _load_categories(self):
        if not self.input_path:
            messagebox.showwarning("尚未選檔", "請先回 Step 1 選擇 Excel 檔")
            return
        try:
            cats = list_categories(self.input_path, ConvertConfig())
        except Exception as e:
            messagebox.showerror("讀檔失敗", str(e))
            return

        # 保留現有已填的樣板（從 entry 收集）
        preserved = {code: e.get() for code, e in self.template_entries.items() if e.get().strip()}

        # 清空 ScrollableFrame
        for w in self.tpl_scroll.winfo_children():
            w.destroy()
        self.template_entries.clear()
        self.template_metadata.clear()

        # 為每個類別生成一列；推測無 BSMI/NCC 的預填 (none)
        auto_filled = 0
        for i, (code, n, sample, hint) in enumerate(cats):
            initial = preserved.get(code, "")
            if not initial and hint not in ("BSMI商品檢驗字號", "NCC型式認證碼"):
                # 推測「依類別需求」（沒命中應施檢驗關鍵字）→ 預填 (none) 提示客戶確認
                initial = "(none)"
                auto_filled += 1
            self._add_category_row(i, code, n, sample, hint, initial)
            self.template_metadata[code] = (n, hint, sample)

        self._refresh_progress()
        messagebox.showinfo(
            "已載入",
            f"檔案中共 {len(cats)} 個商品類別。\n\n"
            f"工具已自動為 {auto_filled} 個推測「無必填屬性」的類別預填 (none)，\n"
            f"請逐一確認 — 如某類別其實需要字號，請改寫為「商品檢驗字號-:[您的字號]」等。\n\n"
            f"剩下 {len(cats) - auto_filled} 個推測需要 BSMI / NCC 的類別，需要您填入真實字號才能上架。",
        )

    def _add_category_row(self, idx: int, code: str, count: int, sample: str, hint: str, initial: str = ""):
        row = ctk.CTkFrame(
            self.tpl_scroll,
            fg_color="#f9fafb" if idx % 2 == 0 else "#ffffff",
            corner_radius=0,
        )
        row.pack(fill="x", pady=1)
        # 類別代碼
        ctk.CTkLabel(row, text=code, font=FONT_MONO, width=105, anchor="w").pack(side="left", padx=3, pady=3)
        # 商品數
        ctk.CTkLabel(row, text=f"{count} 筆", font=FONT_BODY, width=55, anchor="center").pack(side="left", padx=3)
        # 推測類型
        color = HINT_TO_PLACEHOLDER.get(hint, ("", DEFAULT_HINT_COLOR))[1]
        ctk.CTkLabel(row, text=hint, font=FONT_BODY, width=130, anchor="w", text_color=color).pack(side="left", padx=3)
        # 代表商品 — 用按鈕，點擊跳出完整品名清單
        sample_short = sample if len(sample) <= 38 else sample[:37] + "…"
        sample_btn = ctk.CTkButton(
            row, text=sample_short, font=FONT_BODY, width=420, height=28,
            anchor="w", fg_color="transparent", text_color="#1d4ed8",
            hover_color="#dbeafe", border_width=0,
            command=lambda c=code: self._show_category_products(c),
        )
        sample_btn.pack(side="left", padx=3)
        # 字號 / 屬性值輸入框
        ph = HINT_TO_PLACEHOLDER.get(hint, ("(none) / 商品檢驗字號-:[您的字號]", ""))[0] or "(none)"
        entry = ctk.CTkEntry(row, font=FONT_BODY, placeholder_text=f"例：{ph}", width=360, height=30)
        entry.pack(side="left", padx=3, pady=2, fill="x", expand=True)
        if initial:
            entry.insert(0, initial)
        entry.bind("<KeyRelease>", lambda e: self._refresh_progress())
        self.template_entries[code] = entry

    def _show_category_products(self, code: str):
        """彈窗顯示該類別所有商品的完整品名"""
        if not self.input_path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.input_path, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            headers = list(next(rows))
            ci = headers.index("類別")
            ni = headers.index("商品名稱")
            names = [str(r[ni]) for r in rows if str(r[ci]) == code and r[ni]]
        except Exception as e:
            messagebox.showerror("讀檔失敗", str(e))
            return

        win = ctk.CTkToplevel(self)
        win.title(f"類別 {code} — 共 {len(names)} 筆商品")
        win.geometry("760x520")
        ctk.CTkLabel(win, text=f"類別 {code}（{len(names)} 筆完整商品名）", font=FONT_SECTION).pack(padx=12, pady=8, anchor="w")
        ctk.CTkLabel(win, text="可複製商品名稱去搜尋對應的 BSMI / NCC 字號", font=FONT_HINT, text_color="gray").pack(padx=12, pady=(0, 4), anchor="w")
        scroll = ctk.CTkScrollableFrame(win)
        scroll.pack(fill="both", expand=True, padx=12, pady=8)
        for i, name in enumerate(names):
            entry = ctk.CTkEntry(scroll, font=FONT_BODY, height=30)
            entry.insert(0, name)
            entry.configure(state="readonly")
            entry.pack(fill="x", pady=1)
        ctk.CTkButton(win, text="關閉", command=win.destroy, font=FONT_BUTTON, width=120).pack(pady=8)

    def _refresh_progress(self):
        total = len(self.template_entries)
        if total == 0:
            self.lbl_tpl_progress.configure(text="尚未載入類別清單", text_color="gray")
            return
        filled = sum(1 for e in self.template_entries.values() if e.get().strip())
        color = "#16a34a" if filled == total else ("#b45309" if filled > 0 else "#dc2626")
        self.lbl_tpl_progress.configure(text=f"已填樣板：{filled} / {total}", text_color=color)

    def _fill_placeholders(self):
        if not self.template_entries:
            self._load_categories()
            if not self.template_entries:
                return
        bsmi_count = ncc_count = none_count = 0
        for code, entry in self.template_entries.items():
            n, hint, sample = self.template_metadata.get(code, (0, "", ""))
            if "NCC" in hint:
                tpl = "NCC型式認證碼-:CCAB12LP3210T0"
                ncc_count += 1
            elif "BSMI" in hint or "檢驗" in hint:
                tpl = "商品檢驗字號-:R12345"
                bsmi_count += 1
            else:
                # 推測無必填屬性 → 填 (none)，不是 BSMI 占位符
                tpl = "(none)"
                none_count += 1
            entry.delete(0, "end")
            entry.insert(0, tpl)
        self._refresh_progress()
        messagebox.showinfo(
            "已填入測試占位符",
            f"已填入：\n"
            f"  • BSMI 占位符：{bsmi_count} 個類別\n"
            f"  • NCC  占位符：{ncc_count} 個類別\n"
            f"  • (none)（推測無必填）：{none_count} 個類別\n\n"
            "⚠️ 占位符僅能通過匯入驗證，實際上架前請替換為真實字號。\n"
            "⚠️ 工具可能誤判類別 — 若 Yahoo 工具回報某類別錯誤，\n"
            "    請手動回到表格修正該行（例如改為「適用電壓-:110V」）。",
        )

    def _save_templates_json(self):
        if not self.template_entries:
            messagebox.showwarning("尚無資料", "請先載入檔案類別清單")
            return
        path = filedialog.asksaveasfilename(
            title="儲存樣板",
            defaultextension=".json",
            filetypes=[("JSON 樣板檔", "*.json")],
            initialfile="yahoo_templates.json",
        )
        if not path:
            return
        data = {code: e.get() for code, e in self.template_entries.items()}
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        messagebox.showinfo("已儲存", f"樣板已儲存到：\n{path}\n\n下次匯入新檔案時可用「📥 載入既有樣板」直接套用。")

    def _load_templates_json(self):
        path = filedialog.askopenfilename(
            title="載入樣板",
            filetypes=[("JSON 樣板檔", "*.json"), ("所有檔案", "*.*")],
        )
        if not path:
            return
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("讀檔失敗", str(e))
            return
        if not self.template_entries:
            messagebox.showwarning("尚未載入類別", "請先按「📂 載入檔案類別」")
            return
        applied = 0
        for code, tpl in data.items():
            if code in self.template_entries:
                self.template_entries[code].delete(0, "end")
                self.template_entries[code].insert(0, tpl)
                applied += 1
        self._refresh_progress()
        messagebox.showinfo(
            "已套用",
            f"從 {len(data)} 個樣板中，{applied} 個對應到此檔案的類別。\n"
            f"未對應的 {len(data) - applied} 個樣板（檔案沒這些類別）已忽略。",
        )

    def _parse_templates(self) -> dict:
        """從表格 entries 收集 {類別代碼: 樣板字串}"""
        return {code: e.get().strip() for code, e in self.template_entries.items() if e.get().strip()}

    def _guess_total(self) -> int:
        """估算檔案總商品數（從類別 metadata 累加）"""
        return sum(n for n, _, _ in self.template_metadata.values()) if self.template_metadata else 0

    def _run_convert(self):
        if not self.input_path:
            messagebox.showwarning("尚未選檔", "請先回 Step 1 選擇 Excel 檔")
            return
        if not self.legal_var.get():
            messagebox.showwarning("須勾選同意", "請勾選「我了解 BSMI 判定責任」後再轉檔")
            return

        # 樣板填寫狀況檢查
        templates = self._parse_templates()
        total_categories = len(self.template_entries)
        filled = len(templates)

        if total_categories == 0:
            messagebox.showwarning(
                "尚未載入類別",
                "請先按「📂 載入檔案類別」載入這個檔案的類別清單，再決定要不要填字號。",
            )
            return

        if filled == 0:
            proceed = messagebox.askyesno(
                "尚未填寫任何字號",
                f"您尚未填寫任何字號樣板（{total_categories} 個類別都空白）。\n\n"
                f"如果繼續轉檔：\n"
                f"  • 工具仍會替您修正運費 / 收款方式等格式問題\n"
                f"  • 但全部 {self._guess_total()} 筆商品會分流到「需人工處理」（無法直接上架）\n\n"
                f"建議：\n"
                f"  • 想直接上架 → 按「取消」回去填字號\n"
                f"  • 只想看修正效果 → 按「是」繼續\n\n"
                f"要繼續轉檔嗎？",
                icon="warning",
            )
            if not proceed:
                return
        elif filled < total_categories:
            proceed = messagebox.askyesno(
                "部分類別尚未填字號",
                f"已填 {filled} / {total_categories} 個類別。\n\n"
                f"未填字號的 {total_categories - filled} 個類別下的商品，會分流到「需人工處理」。\n"
                f"您可以先轉檔看部分結果，之後再回來補上其他類別。\n\n"
                f"要繼續轉檔嗎？",
            )
            if not proceed:
                return

        # 解析參數
        try:
            cfg = ConvertConfig(
                fee_711=int(self.fee_vars["7-11 取貨付款"].get()),
                fee_familymart=int(self.fee_vars["全家取貨付款"].get()),
                fee_hilife=int(self.fee_vars["萊爾富取貨付款"].get()),
                fee_post_cod=int(self.fee_vars["郵局貨到付款"].get()),
                fee_mail=int(self.fee_vars["郵寄掛號"].get()),
                fee_home=int(self.fee_vars["宅配"].get()),
                use_store_shipping=self.use_store_var.get(),
                category_attr_templates=templates,
            )
        except ValueError as e:
            messagebox.showerror("參數錯誤", f"運費必須是數字\n{e}")
            return

        def work():
            try:
                def cb(i, total):
                    self.progress.set(i / total)
                    self.lbl_progress.configure(text=f"處理中... {i} / {total}")
                    self.update_idletasks()

                self.result = convert(self.input_path, cfg, progress_cb=cb)
                self.lbl_progress.configure(text="✅ 轉檔完成")
                self._show_result()
                self.tabview.set("Step 3：預覽結果")
            except Exception as e:
                messagebox.showerror("轉檔失敗", str(e))
                self.lbl_progress.configure(text=f"❌ 失敗: {e}", text_color="red")

        threading.Thread(target=work, daemon=True).start()

    # ---------- Step 3 ----------
    def _build_step3(self):
        frame = self.tab_step3
        warn_banner = ctk.CTkFrame(frame, fg_color="#fef2f2", border_color="#dc2626", border_width=2)
        warn_banner.pack(fill="x", padx=10, pady=(6, 4))
        ctk.CTkLabel(
            warn_banner,
            text="⚠️ 通過格式驗證 ≠ 可實際上架。占位符（R12345 / CCAB12LP3210T0）必須替換為真實字號方能上架。",
            font=FONT_BUTTON, text_color="#dc2626", wraplength=1100,
        ).pack(padx=8, pady=6)

        self.stat_frame = ctk.CTkFrame(frame)
        self.stat_frame.pack(fill="x", padx=10, pady=4)
        self.lbl_stats = ctk.CTkLabel(self.stat_frame, text="尚未轉檔", font=FONT_SECTION)
        self.lbl_stats.pack(padx=8, pady=6)

        # ttk.Treeview 字體用 style 設定
        style = ttk.Style()
        style.configure("Treeview", font=("Microsoft JhengHei", 11), rowheight=26)
        style.configure("Treeview.Heading", font=("Microsoft JhengHei", 12, "bold"))

        self.inner_tab = ctk.CTkTabview(frame, height=460)
        self.inner_tab.pack(fill="both", expand=True, padx=10, pady=4)
        tab_specs = [
            ("✅ 可上架", "tree_uploadable"),
            ("⚠️ 需人工處理", "tree_review"),
            ("📋 修正紀錄", "tree_log"),
        ]
        for tab_name, attr in tab_specs:
            self.inner_tab.add(tab_name)
            tab = self.inner_tab.tab(tab_name)
            tree = ttk.Treeview(tab, show="headings")
            tree.pack(fill="both", expand=True, side="left")
            sb = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
            sb.pack(side="right", fill="y")
            tree.configure(yscrollcommand=sb.set)
            setattr(self, attr, tree)

        btn_box = ctk.CTkFrame(frame, fg_color="transparent")
        btn_box.pack(fill="x", padx=10, pady=6)
        ctk.CTkButton(btn_box, text="← 修改參數", command=lambda: self.tabview.set("Step 2：設定參數"), width=140, height=40, font=FONT_BUTTON).pack(side="left")
        ctk.CTkButton(
            btn_box, text="💾 匯出 Excel...", command=self._export,
            width=220, height=44, font=FONT_BIG_BUTTON,
        ).pack(side="right")

    def _show_result(self):
        s = self.result.stats
        text = (
            f"總計 {s['total']} 筆 ｜ "
            f"✅ 可上架 {s['uploadable']} 筆 ｜ "
            f"⚠️ 需人工處理 {len(self.result.needs_review)} 筆 "
            f"(屬性欄位1空 {s['review_attr']} / 數量空白 {s['review_qty']} / 其他 {s['review_other']})  ｜ "
            f"📋 自動修正 {s['fixes']} 個欄位"
        )
        self.lbl_stats.configure(text=text)

        # 如果可上架=0 但有修正紀錄，顯示安撫訊息
        if hasattr(self, "info_banner") and self.info_banner is not None:
            self.info_banner.destroy()
            self.info_banner = None
        if s["uploadable"] == 0 and s["fixes"] > 0:
            self.info_banner = ctk.CTkFrame(self.stat_frame, fg_color="#dcfce7", border_color="#16a34a", border_width=2)
            self.info_banner.pack(fill="x", padx=8, pady=(4, 6))
            ctk.CTkLabel(
                self.info_banner,
                text=(
                    f"ℹ️ 您尚未填寫字號樣板，但工具已替您完成 {s['fixes']} 個金物流欄位修正。\n"
                    f"   「修正紀錄」工作表可看完整明細。下一步：填字號樣板再轉一次，即可獲得可上架檔案。"
                ),
                font=FONT_LABEL, text_color="#15803d", justify="left",
            ).pack(padx=10, pady=8, anchor="w")

        # 三個 tree（限制每個 tab 最多顯示 300 筆，避免 GUI 卡）
        self._fill_tree(
            self.tree_uploadable,
            ["列號", "商品名稱", "類別代碼", "定價"],
            [(i+1, self._get(r, "商品名稱"), self._get(r, "類別"), self._get(r, "定價")) for i, r in enumerate(self.result.uploadable[:300])],
        )
        review_headers = ["列號", "商品名稱", "類別代碼", "未通過原因"]
        review_data = []
        for i, r in enumerate(self.result.needs_review[:300]):
            review_data.append((i+1, self._get(r, "商品名稱"), self._get(r, "類別"), r[-1]))
        self._fill_tree(self.tree_review, review_headers, review_data)

        log_data = [(e.row_index, e.product_name[:40], e.field, str(e.old_value)[:30], str(e.new_value)[:30], e.rule) for e in self.result.fix_log[:500]]
        self._fill_tree(self.tree_log, ["原列", "商品名稱", "欄位", "原值", "新值", "規則"], log_data)

    def _get(self, row, field):
        try:
            i = self.result.headers.index(field)
            return str(row[i]) if row[i] is not None else ""
        except ValueError:
            return ""

    def _fill_tree(self, tree, columns, rows):
        tree.delete(*tree.get_children())
        tree["columns"] = columns
        for c in columns:
            tree.heading(c, text=c)
            tree.column(c, width=180 if "名稱" in c else 100, anchor="w")
        for r in rows:
            tree.insert("", "end", values=r)

    def _export(self):
        if not self.result:
            return
        # 二次確認
        confirm = messagebox.askyesno(
            "上架前最後提醒",
            "即將匯出 Excel 檔。在您將此檔案上架到 Yahoo 之前，請確認：\n\n"
            "✓ 所有應施檢驗商品已填入真實有效的 BSMI 字號\n"
            "✓ 所有通訊類商品已填入真實有效的 NCC 認證碼\n"
            "✓ 已替換掉測試占位符（R12345、CCAB12LP3210T0）\n\n"
            "未確認上述內容直接上架，依商品檢驗法可處 20–200 萬罰鍰。\n"
            "您確認已了解此風險，要繼續匯出嗎？",
            icon="warning",
        )
        if not confirm:
            return
        out = filedialog.asksaveasfilename(
            title="儲存轉檔結果",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="yahoo_ready.xlsx",
        )
        if not out:
            return
        try:
            write_output(self.result, out)
            messagebox.showinfo(
                "匯出完成",
                f"已儲存到：\n{out}\n\n"
                "包含三個工作表：\n"
                "  ✅ 可上架商品（第一個工作表 → 直接給 Yahoo 大量刊登助手匯入）\n"
                "  ⚠️ 需人工處理（輔助參考用）\n"
                "  📋 修正紀錄（自動修正明細）\n\n"
                "⚠️ 上架前請替換測試占位符（R12345 / CCAB12LP3210T0）為真實字號。",
            )
        except Exception as e:
            messagebox.showerror("匯出失敗", str(e))


if __name__ == "__main__":
    App().mainloop()
