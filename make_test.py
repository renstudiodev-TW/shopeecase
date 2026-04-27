import openpyxl, os, shutil, sys, io
from PIL import Image
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

src = r"C:\RenStudio\case\testcase\shopee\拍賣蝦皮拍賣商品_匯出商品資料_20260421_024920.xlsx"
dst = r"C:\RenStudio\case\testcase\shopee\test_v1.xlsx"
photos = r"C:\RenStudio\case\testcase\shopee\photos"
os.makedirs(photos, exist_ok=True)

# 複製原檔
shutil.copy(src, dst)

# 開啟複製檔，只保留前 2 筆
wb = openpyxl.load_workbook(dst)
ws = wb.active
print(f"原始列數: {ws.max_row}")

# 從第 4 列開始全部刪掉（保留 header + 2 筆資料）
ws.delete_rows(4, ws.max_row)
print(f"裁剪後列數: {ws.max_row}")

# 找出圖片欄位
headers = [c.value for c in ws[1]]
img_cols = [i+1 for i, h in enumerate(headers) if h and str(h).startswith("圖片")]
print(f"圖片欄位 col index: {img_cols[:5]} ...")

# 收集前 2 筆所有出現過的圖片檔名
img_names = set()
for row in [2, 3]:
    for c in img_cols:
        v = ws.cell(row=row, column=c).value
        if v:
            img_names.add(str(v))

print(f"前 2 筆共需要 {len(img_names)} 張圖")

# 在 photos 資料夾建立同名假圖（800x800 白底）
for name in img_names:
    path = os.path.join(photos, name)
    img = Image.new("RGB", (800, 800), "white")
    img.save(path, "JPEG", quality=85)

print(f"✅ 已在 {photos} 建立 {len(img_names)} 張 800x800 假圖")

wb.save(dst)
print(f"✅ 測試檔已儲存: {dst}")
