import openpyxl, sys, io, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

src = r"C:\RenStudio\case\testcase\shopee\test_v1.xlsx"
dst = r"C:\RenStudio\case\testcase\shopee\test_v2.xlsx"
shutil.copy(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb.active
headers = [c.value for c in ws[1]]
def col(name): return headers.index(name) + 1

# 收款方式 → 對應的運送方式
mapping = {
    "接受輕鬆付7-11取貨付款": "7-11取貨付款",
    "接受輕鬆付萊爾富取貨付款": "萊爾富取貨付款",
    "接受輕鬆付全家取貨付款": "全家取貨付款",
    "接受輕鬆付郵局貨到付款": "郵局貨到付款",
}
# 接受現金/信用卡 → 至少一種可預付（用郵寄掛號）
prepay_triggers = ["接受輕鬆付現金付款", "接受輕鬆付信用卡一次付清"]
prepay_target = "郵寄掛號"

fixed = 0
for row in range(2, ws.max_row + 1):
    for pay_col, ship_col in mapping.items():
        if str(ws.cell(row=row, column=col(pay_col)).value).lower() == "yes":
            cell = ws.cell(row=row, column=col(ship_col))
            if str(cell.value).lower() != "yes":
                cell.value = "yes"
                fixed += 1
                print(f"  row {row}: {ship_col} → yes (因為 {pay_col} = yes)")
    # 預付類
    if any(str(ws.cell(row=row, column=col(t)).value).lower() == "yes" for t in prepay_triggers):
        cell = ws.cell(row=row, column=col(prepay_target))
        if str(cell.value).lower() != "yes":
            cell.value = "yes"
            fixed += 1
            print(f"  row {row}: {prepay_target} → yes (因為接受現金/信用卡)")

wb.save(dst)
print(f"\n✅ 共修正 {fixed} 個欄位，存檔到 test_v2.xlsx")
