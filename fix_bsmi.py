import openpyxl, sys, io, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

src = r"C:\RenStudio\case\testcase\shopee\test_v2.xlsx"
dst = r"C:\RenStudio\case\testcase\shopee\test_v3.xlsx"
shutil.copy(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb.active

# 找一個合理位置插入「商品檢驗字號」欄
# 通常放在「商品新舊」之後比較合理（屬於商品基本資料）
headers = [c.value for c in ws[1]]

# 看「商品新舊」在哪
target = "商品新舊"
if target in headers:
    insert_at = headers.index(target) + 2  # +1 是因為 openpyxl 1-based, 再 +1 是要插在它後面
    ws.insert_cols(insert_at)
    ws.cell(row=1, column=insert_at).value = "商品檢驗字號"
    # 填值（測試用 - 全部標示為無需檢驗）
    placeholder = "商品無需檢驗"
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=insert_at).value = placeholder
    print(f"✅ 已在第 {insert_at} 欄（{target} 後面）插入「商品檢驗字號」欄")
    print(f"✅ 全部 {ws.max_row - 1} 筆填入「{placeholder}」（僅測試用）")
else:
    print(f"❌ 找不到「{target}」欄位")

wb.save(dst)
print(f"✅ 存檔 → {dst}")

# 驗證
wb2 = openpyxl.load_workbook(dst, read_only=True)
ws2 = wb2.active
new_headers = [c.value for c in next(ws2.iter_rows(values_only=False))]
print(f"\n新檔案總欄位數: {len(new_headers)}")
print(f"商品檢驗字號 在第 {new_headers.index('商品檢驗字號')+1} 欄")
